# 게시물
from openpyxl import Workbook, load_workbook

wb = Workbook()
ws = wb.active
ws.title = "Articles"
ws["F1"] = "rowCount"
ws["G1"] = 4

# ws["A2"] = "번호"
# ws["B2"] = "제목"
# ws["C2"] = "내용"
# ws["D2"] = "작성자"

ws["A1"] = "1"
ws["B1"] = "소니의 축구교실"
ws["C1"] = "소니의 축구 강좌"
ws["D1"] = "sony7"

ws["A2"] = "2"
ws["B2"] = "류뚱의 야구교실"
ws["C2"] = "류뚱의 야구 강좌"
ws["D2"] = "ryu99"

ws["A3"] = "3"
ws["B3"] = "길동의 도술교술"
ws["C3"] = "길동의 도술 강좌"
ws["D3"] = "hong123"

wb.save("ArticleWB.xlsx")


# Save posts to Xlsx
def writeArticleToXl(article):
    wb = load_workbook("ArticleWB.xlsx")
    ws = wb["Articles"]
    rowCnt = ws["G1"].value

    ws["A" + str(rowCnt + 1)] = article["번호"]
    ws["B" + str(rowCnt + 1)] = article["제목"]
    ws["C" + str(rowCnt + 1)] = article["내용"]
    ws["D" + str(rowCnt + 1)] = article["작성자"]

    ws["G1"] = rowCnt + 1
    wb.save("ArticleWB.xlsx")


# Read posts from Xlsx
def readArticleFrXl():
    wb = load_workbook("ArticleWB.xlsx")
    ws = wb["Articles"]

    articles = []
    for r in ws.rows:
        row_index = r[0].row
        title = r[1].value
        body = r[2].value
        user = r[3].value
        article = {"번호": row_index, "제목": title, "내용": body, "작성자": user}
        articles.append(article)

    return articles


# Update posts in Xlsx
def updateArticleToXL(num, title, body):
    wb = load_workbook("ArticleWB.xlsx")
    ws = wb["Articles"]

    for r in ws.rows:
        row_index = r[0].value
        if row_index == num:
            r[1].value = title
            r[2].value = body
            return 1
            wb.save("ArticleWB.xlsx")


# Delete posts from Xlsx
def deleteArticleFrXL(num):
    wb = load_workbook("ArticleWB.xlsx")
    ws = wb["Articles"]

    for r in ws.rows:
        row_index = r[0].value
        if row_index == num:
            ws.delete_rows(r)
            return 1
            wb.save("ArticleWB.xlsx")


# ==========================================================

# 작성자
wb = load_workbook("ArticleWB.xlsx")
ws2 = wb.create_sheet("Users")
ws2.title = "Users"

ws2["F1"] = "rowCount"
ws2["G1"] = 4

# ws2["A2"] = "번호"
# ws2["B2"] = "아이디"
# ws2["C2"] = "비밀번호"
# ws2["D2"] = "이름"

ws2["A1"] = "1"
ws2["B1"] = "hong123"
ws2["C1"] = "1234"
ws2["D1"] = "홍길동"

ws2["A2"] = "2"
ws2["B2"] = "sony7"
ws2["C2"] = "7777"
ws2["D2"] = "손흥민"

ws2["A3"] = "3"
ws2["B3"] = "ryu99"
ws2["C3"] = "9999"
ws2["D3"] = "류현진"

wb.save("ArticleWB.xlsx")


# New user information
def newUserInfoToXl():
    id_info = input("등록할 아이디를 입력해 주세요 : ")
    pw_info = input("등록할 비밀번호를 입력해 주세요 : ")
    name_info = input("이름을 입력해 주세요 : ")

    new_user = {"아이디": id_info, "비밀번호": pw_info, "이름": name_info}
    return new_user


# Save new user to Xlsx
def addUserToXL(new_user):
    wb = load_workbook("ArticleWB.xlsx")
    ws2 = wb["Users"]
    rowCnt = ws2["G1"].value

    ws2["A" + str(rowCnt + 1)] = rowCnt
    ws2["B" + str(rowCnt + 1)] = new_user["아이디"]
    ws2["C" + str(rowCnt + 1)] = new_user["비밀번호"]
    ws2["D" + str(rowCnt + 1)] = new_user["이름"]

    ws2["G1"] = rowCnt + 1
    wb.save("ArticleWB.xlsx")


def loginUserFrXlsx(user_id, user_pw):
    wb = load_workbook("ArticleWB.xlsx")
    ws2 = wb["Users"]

    for r in ws2.rows:
        row_id = r[1].value
        row_pw = r[2].value
        row_name = r[3].value
        if row_id == user_id:
            if row_pw == user_pw:
                user_dict = {"아이디": row_id, "비밀번호": row_pw, "이름": row_name}
                return user_dict


# Read user from Xlsx
def readUserFrXlsx(user_id, user_pw):
    wb = load_workbook("ArticleWB.xlsx")
    ws2 = wb["Users"]

    for r in ws2.rows:
        row_id = r[1].value
        row_pw = r[2].value
        row_name = r[3].value
        if row_id == user_id:
            if row_pw == user_pw:
                user_dict = {"아이디": row_id, "비밀번호": row_pw, "이름": row_name}
                return user_dict


# Update user to Xlsx
def updateUserToXlsx(new_id, new_pw, name):
    wb = load_workbook("ArticleWB.xlsx")
    ws2 = wb["Users"]

    for r in ws2.rows:
        row_name = r[3].value
        if row_name == name:
            r[0].value = new_id
            r[1].value = new_pw
            wb.save("ArticleWB.xlsx")
            return 1


# Delete user from Xlsx
def deleteUserFrXlsx(del_id):
    wb = load_workbook("ArticleWB.xlsx")
    ws2 = wb["Users"]

    for r in ws2.rows:
        row_id = r[1].value
        if row_id == del_id:
            ws2.delete_rows(r)
            wb.save("ArticleWB.xlsx")
            return 1


# ==========================================================
no = 4


def loginCheck(user_id, user_pw):
    user_dict = readUserFrXlsx(user_id, user_pw)
    if user_dict is None:
        print("없는 아이디입니다")
        return False

    else:
        print("{}님 반갑습니다!".format(user_dict["이름"]))
        return True


def listUser():
    info = readUserFrXlsx(login_id, login_pw)
    print("회원님의 정보 : {}".format(info))


def updateUser():
    name = input("수정할 작성자의 이름을 입력하세요 : ")
    new_id = input("새로운 아이디를 입력하세요 : ")
    new_pw = input("새로운 암호를 입력하세요 : ")
    reply = updateUserToXlsx(new_id, new_pw, name)
    if reply == 1:
        print("작성자 정보가 수정되었습니다.")
    else:
        print("잘못된정보입니다.")


def deleteUser():
    del_id = input("삭제할 아이디를 입력해주세요 : ")
    reply = deleteUserFrXlsx(del_id)
    if reply == 1:
        print("작성자 정보가 삭제되었습니다.")
    else:
        print("잘못된정보입니다.")


def printArticle():
    article = readArticleFrXl()
    print("=========== 게시물 목록 ==============")
    print("번호 : {}".format(article["번호"]))
    print("제목 : {}".format(article["제목"]))
    print("=====================================")


def addArticle():
    global no
    title = input("제목을 입력해주세요")
    body = input("내용을 입력해주세요")
    article = {"번호": no, "제목": title, "내용": body, "작성자": login_id}
    writeArticleToXl(article)
    no += 1


def updateArticle():
    num = int(input("게시물 번호를 입력해주세요 : "))
    title = input("수정 제목 : ")
    body = input("수정 내용 : ")
    target = updateArticleToXL(num, title, body)

    if target == 1:
        print("수정이 완료되었습니다.")

    else:
        print("없는 게시물입니다.")


def deleteArticle():
    num = int(input("게시물 번호를 입력해주세요 : "))
    target = deleteArticleFrXL(num)

    if target == 1:
        print("삭제가 완료되었습니다.")

    else:
        print("없는 게시물입니다.")


def detailArticle():
    articles = readArticleFrXl()
    for article in articles:
        print("=========  게시물 목록  ==========")
        print("번호 : {}".format(article["번호"]))
        print("제목 : {}".format(article["제목"]))
        print("내용 : {}".format(article["내용"]))
        print("작성자 : {}".format(article["작성자"]))
        print("=================================")

    if articles is None:
        print("게시물이 없습니다.")


def printHelp():
    print("add : 게시물 추가")
    print("list : 게시물 조회")
    print("add : 게시물 추가")
    print("update : 게시물 수정")
    print("delete : 게시물 삭제")
    print("add user : 작성자 추가")
    print("print user : 작성자 정보 조회")
    print("update user : 작성자 수정")
    print("delete_user : 작성자 삭제")


# ===========================================================

login_id = input("아이디를 입력해주세요 : ")
login_pw = input("비밀번호를 입력해주세요 : ")

loginResult = loginCheck(login_id, login_pw)

if loginResult:
    while True:
        cmd = input("명령어를 입력해 주세요 :")
        if cmd == "exit":
            print("프로그램을 종료합니다.")
            break

        elif cmd == "help":
            printHelp()

        elif cmd == "list":
            printArticle()

        elif cmd == "add":
            addArticle()

        elif cmd == "update":
            updateArticle()

        elif cmd == "delete":
            deleteArticle()

        elif cmd == "detail":
            detailArticle()

        elif cmd == "add user":
            user_info = newUserInfoToXl()
            addUserToXL(user_info)

        elif cmd == "print user":
            listUser()

        elif cmd == "update user":
            updateUser()

        elif cmd == "delete_user":
            deleteUser()

# ===========================================================
